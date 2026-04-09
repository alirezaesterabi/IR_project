"""
Run queries_type_{1-6}.xlsx against all retrieval models, fuse with RRF,
cross-reference against expected_doc_ids ground truth, and output:
  - One CSV per query type (in results/)
  - One combined Excel workbook with all types (in results/)

Usage
-----
    python scripts/test_script.py

Outputs (in results/)
---------------------
    type_1_fused_all_retrievers_audit_{timestamp}.csv
    type_2_fused_all_retrievers_audit_{timestamp}.csv
    ...
    type_6_fused_all_retrievers_audit_{timestamp}.csv
    all_types_fused_audit_{timestamp}.xlsx
        Sheets: Type1_RRF_Fused_Results, Type1_Ground_Truth_Eval, Type1_Summary,
                Type2_RRF_Fused_Results, ..., Overall_Summary
"""

import ast
import csv
import json
import os
import sys
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

# Timestamp suffix for all output files: YYYYMMDD_HHMMSS
RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")


# ── Locate project root ────────────────────────────────────────────────────

def find_root() -> Path:
    marker = "data/raw_data"
    for key in ("IR_PROJECT_ROOT", "VSCODE_WORKSPACE_FOLDER"):
        v = os.environ.get(key)
        if v and (Path(v) / marker).exists():
            return Path(v)
    for p in [Path.cwd()] + list(Path.cwd().parents):
        if (p / marker).exists():
            return p
    raise FileNotFoundError("Cannot find project root.")


ROOT = find_root()
sys.path.insert(0, str(ROOT))

TOP_K = 100  # number of results per query


# ── Helpers ─────────────────────────────────────────────────────────────────

def results_to_ranked_list(
    results: list[tuple[str, float]]
) -> list[tuple[str, int]]:
    """Convert retriever output [(doc_id, score), ...] to [(doc_id, rank), ...]."""
    return [(doc_id, rank) for rank, (doc_id, _score) in enumerate(results, start=1)]


# ── Load queries ────────────────────────────────────────────────────────────

def load_queries(xlsx_path: Path) -> list[dict]:
    """
    Load queries from Excel. Each row has query_id, query_texts,
    and expected_doc_ids (both JSON-stringified lists).
    """
    df = pd.read_excel(xlsx_path)
    queries = []
    for _, row in df.iterrows():
        qid = str(row["query_id"])
        raw = row["query_texts"]
        try:
            texts = ast.literal_eval(raw)
        except (ValueError, SyntaxError):
            texts = [str(raw)]
        if isinstance(texts, str):
            texts = [texts]

        raw_expected = row.get("expected_doc_ids", "[]")
        try:
            expected = ast.literal_eval(str(raw_expected))
        except (ValueError, SyntaxError):
            expected = [str(raw_expected)] if pd.notna(raw_expected) else []
        if isinstance(expected, str):
            expected = [expected]

        queries.append({
            "query_id": qid,
            "query_texts": texts,
            "expected_doc_ids": expected,
        })
    return queries


# ── Excel formatting helpers ───────────────────────────────────────────────

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
EXPECTED_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
MULTI_HIT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FOUND_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MISSED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def style_header(ws, ncols):
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def auto_width(ws, dataframe):
    for col_idx, col_name in enumerate(dataframe.columns, start=1):
        max_len = max(len(str(col_name)), dataframe[col_name].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)


# ── Process one query type (returns dataframes, no file I/O) ──────────────

def process_query_type(
    query_type: int,
    queries: list[dict],
    all_retriever_results: dict[str, dict],
    retriever_keys: list[str],
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    Fuse results, evaluate against ground truth.
    Returns (fused_df, eval_df, summary_stats).
    """
    from src.fusion.rrf import ReciprocalRankFusion

    rrf = ReciprocalRankFusion(k=60)

    expected_lookup = {}
    for q in queries:
        expected_lookup[q["query_id"]] = set(q["expected_doc_ids"])

    all_qids = sorted(set(
        qid
        for results in all_retriever_results.values()
        for qid in results
    ))

    # Build fused results with audit columns + ground truth flag
    fused_rows = []
    for qid in all_qids:
        rank_lookups = {}
        ranked_lists = []
        for label in retriever_keys:
            pairs = all_retriever_results[label].get(qid, [])
            ranked_lists.append(pairs)
            rank_lookups[label] = {doc_id: rank for doc_id, rank in pairs}

        fused = rrf.fuse(*ranked_lists)
        expected_ids = expected_lookup.get(qid, set())

        for fused_rank, (doc_id, score) in enumerate(fused, start=1):
            row = {
                "query_id": qid,
                "doc_id": doc_id,
                "fused_rank": fused_rank,
                "rrf_score": round(score, 6),
                "is_expected": "YES" if doc_id in expected_ids else "",
            }
            for label in retriever_keys:
                row[f"rank_{label}"] = rank_lookups[label].get(doc_id, None)
            fused_rows.append(row)

    df = pd.DataFrame(fused_rows)

    # Ground truth evaluation per query
    eval_rows = []
    for q in queries:
        qid = q["query_id"]
        expected_ids = set(q["expected_doc_ids"])
        qdf = df[df["query_id"] == qid]

        for exp_id in expected_ids:
            match = qdf[qdf["doc_id"] == exp_id]
            found = len(match) > 0
            fused_rank = int(match.iloc[0]["fused_rank"]) if found else None

            retriever_hits = {}
            for label in retriever_keys:
                rank_val = match.iloc[0][f"rank_{label}"] if found else None
                retriever_hits[label] = int(rank_val) if pd.notna(rank_val) else None

            eval_rows.append({
                "query_id": qid,
                "query_texts": str(q["query_texts"]),
                "expected_doc_id": exp_id,
                "found_in_results": "YES" if found else "NO",
                "fused_rank": fused_rank,
                **{f"rank_{r}": retriever_hits[r] for r in retriever_keys},
            })

    eval_df = pd.DataFrame(eval_rows)

    found_count = int((eval_df["found_in_results"] == "YES").sum())
    total_count = len(eval_df)
    recall = found_count / total_count * 100 if total_count > 0 else 0
    found_ranks = eval_df.loc[eval_df["found_in_results"] == "YES", "fused_rank"]

    # Print evaluation
    print(f"\n  Ground truth evaluation:")
    print(f"    Expected docs found: {found_count}/{total_count} ({recall:.1f}%)")
    if len(found_ranks) > 0:
        print(f"    Fused rank — min: {found_ranks.min()}, "
              f"max: {found_ranks.max()}, median: {found_ranks.median():.0f}")
    for label in retriever_keys:
        hits = eval_df[f"rank_{label}"].notna().sum()
        print(f"    {label:<15} found: {hits}/{total_count}")
    missed = eval_df[eval_df["found_in_results"] == "NO"]
    if len(missed) > 0:
        print(f"\n    Missed queries ({len(missed)}):")
        for _, row in missed.iterrows():
            print(f"      {row['query_id']}: expected {row['expected_doc_id']}")

    summary = {
        "query_type": query_type,
        "total_queries": total_count,
        "found": found_count,
        "missed": total_count - found_count,
        "recall_pct": recall,
        "median_rank": float(found_ranks.median()) if len(found_ranks) > 0 else None,
        "min_rank": int(found_ranks.min()) if len(found_ranks) > 0 else None,
        "max_rank": int(found_ranks.max()) if len(found_ranks) > 0 else None,
        "per_retriever": {
            r: int(eval_df[f"rank_{r}"].notna().sum()) for r in retriever_keys
        },
    }

    return df, eval_df, summary


# ── Write sheets for one query type into an existing ExcelWriter ──────────

def write_type_sheets(
    writer: pd.ExcelWriter,
    query_type: int,
    fused_df: pd.DataFrame,
    eval_df: pd.DataFrame,
    summary: dict,
    retriever_keys: list[str],
):
    prefix = f"Type{query_type}"

    # Sheet: RRF Fused Results
    sheet_rrf = f"{prefix}_RRF_Fused_Results"
    fused_df.to_excel(writer, index=False, sheet_name=sheet_rrf)
    ws1 = writer.sheets[sheet_rrf]
    style_header(ws1, len(fused_df.columns))
    auto_width(ws1, fused_df)

    rank_cols = [fused_df.columns.get_loc(f"rank_{r}") + 1 for r in retriever_keys]
    is_expected_col = fused_df.columns.get_loc("is_expected") + 1

    for row_idx in range(2, len(fused_df) + 2):
        is_exp = ws1.cell(row=row_idx, column=is_expected_col).value
        non_empty = sum(1 for c in rank_cols if ws1.cell(row=row_idx, column=c + 1).value is not None)

        if is_exp == "YES":
            for col_idx in range(1, len(fused_df.columns) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = EXPECTED_FILL
        elif non_empty >= 2:
            for col_idx in range(1, len(fused_df.columns) + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = MULTI_HIT_FILL

    ws1.freeze_panes = "C2"
    ws1.auto_filter.ref = ws1.dimensions

    # Sheet: Ground Truth Evaluation
    sheet_eval = f"{prefix}_Ground_Truth_Eval"
    eval_df.to_excel(writer, index=False, sheet_name=sheet_eval)
    ws2 = writer.sheets[sheet_eval]
    style_header(ws2, len(eval_df.columns))
    auto_width(ws2, eval_df)

    found_col = eval_df.columns.get_loc("found_in_results") + 1
    for row_idx in range(2, len(eval_df) + 2):
        val = ws2.cell(row=row_idx, column=found_col).value
        fill = FOUND_FILL if val == "YES" else MISSED_FILL
        for col_idx in range(1, len(eval_df.columns) + 1):
            ws2.cell(row=row_idx, column=col_idx).fill = fill

    ws2.freeze_panes = "C2"
    ws2.auto_filter.ref = ws2.dimensions

    # Sheet: Summary
    sheet_sum = f"{prefix}_Summary"
    summary_data = {
        "Metric": [
            "Query type",
            "Total queries",
            "Expected docs found",
            "Expected docs missed",
            "Recall (%)",
            "Best fused rank (of found)",
            "Worst fused rank (of found)",
            "Median fused rank (of found)",
        ] + [f"Recall — {r}" for r in retriever_keys],
        "Value": [
            f"Type {summary['query_type']}",
            summary["total_queries"],
            summary["found"],
            summary["missed"],
            f"{summary['recall_pct']:.1f}%",
            summary["min_rank"] if summary["min_rank"] is not None else "N/A",
            summary["max_rank"] if summary["max_rank"] is not None else "N/A",
            f"{summary['median_rank']:.0f}" if summary["median_rank"] is not None else "N/A",
        ] + [
            f"{summary['per_retriever'][r]}/{summary['total_queries']}"
            for r in retriever_keys
        ],
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, index=False, sheet_name=sheet_sum)
    ws3 = writer.sheets[sheet_sum]
    style_header(ws3, len(summary_df.columns))
    auto_width(ws3, summary_df)


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    models_dir = ROOT / "models"
    chroma_dir = ROOT / "chroma_db"
    queries_dir = ROOT / "data" / "potential_queries"
    output_dir = ROOT / "results"
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = RUN_TIMESTAMP

    query_types = [1, 2, 3, 4, 5, 6]

    print("=" * 60)
    print("  Test Script — all query types against all retrievers")
    print("=" * 60)
    print(f"  Query types : {query_types}")
    print(f"  Models      : {models_dir}")
    print(f"  Output      : {output_dir}")
    print(f"  Timestamp   : {ts}")
    print()

    # ── Load all retrievers once ──────────────────────────────────────

    from src.retrieval.classical_ir import BM25Retriever, TFIDFRetriever, IdentifierRetriever
    from src.preprocessing.text_processing import TextProcessor
    from src.retrieval.dense_retriever import DenseRetriever
    from src.retrieval.dense_config import DENSE_MODELS, model_file_suffix, compute_run_tag
    import chromadb

    print("Loading models...")
    tp = TextProcessor()

    bm25 = BM25Retriever()
    bm25.load(models_dir / "bm25")

    tfidf = TFIDFRetriever()
    tfidf.load(models_dir / "tfidf")

    ident = IdentifierRetriever()
    ident.load(models_dir / "identifier" / "index.pkl")

    # Discover available ChromaDB collections and build DenseRetrievers
    client = chromadb.PersistentClient(path=str(chroma_dir))
    existing_colls = [c.name for c in client.list_collections()]

    dense_retrievers: dict[str, DenseRetriever] = {}
    for model_key in ["minilm", "bge-m3"]:
        suffix = model_file_suffix(model_key)
        run_tag = compute_run_tag(None)
        coll_name = f"opensanctions_{suffix}_{run_tag}"
        if coll_name not in existing_colls:
            matching = sorted([n for n in existing_colls if n.startswith(f"opensanctions_{suffix}_")])
            coll_name = matching[-1] if matching else None
        if coll_name:
            # Extract run_tag from the discovered collection name
            prefix = f"opensanctions_{suffix}_"
            discovered_run_tag = coll_name[len(prefix):]
            print(f"  {model_key}: collection {coll_name}")
            dense_retrievers[model_key] = DenseRetriever(
                model_key=model_key,
                run_tag=discovered_run_tag,
                chroma_dir=chroma_dir,
                models_dir=models_dir,
            )
        else:
            print(f"  WARNING: No collection found for {model_key}")

    retriever_keys = ["bm25", "tfidf", "identifier", "dense_minilm", "dense_bge_m3"]
    print("\nAll models loaded.\n")

    # ── Run all retrievers for a set of queries ───────────────────────

    def run_all_retrievers(queries: list[dict]) -> dict[str, dict]:
        results = {}

        # BM25
        r = {}
        for q in queries:
            combined = " ".join(q["query_texts"])
            tokens = tp.tokenize_name(combined)
            hits = bm25.search(tokens, k=TOP_K)
            r[q["query_id"]] = results_to_ranked_list(hits)
        results["bm25"] = r

        # TF-IDF
        r = {}
        for q in queries:
            combined = " ".join(q["query_texts"])
            normalised = tp.normalize(combined)
            hits = tfidf.search(normalised, k=TOP_K)
            r[q["query_id"]] = results_to_ranked_list(hits)
        results["tfidf"] = r

        # Identifier
        r = {}
        for q in queries:
            seen = set()
            hits = []
            for text in q["query_texts"]:
                for doc_id, score in ident.search(text):
                    if doc_id not in seen:
                        seen.add(doc_id)
                        hits.append((doc_id, score))
            r[q["query_id"]] = results_to_ranked_list(hits)
        results["identifier"] = r

        # Dense models (via DenseRetriever)
        for model_key, label in [("minilm", "dense_minilm"), ("bge-m3", "dense_bge_m3")]:
            if model_key not in dense_retrievers:
                results[label] = {}
                continue
            retriever = dense_retrievers[model_key]
            r = {}
            for q in queries:
                query_text = " ".join(q["query_texts"])
                hits = retriever.search(query_text, k=TOP_K, include_metadata=False)
                ranked = [(doc_id, score) for doc_id, score, _meta in hits]
                r[q["query_id"]] = results_to_ranked_list(ranked)
            results[label] = r

        return results

    # ── Process each query type ───────────────────────────────────────

    all_summaries = []
    all_type_data = []  # (query_type, fused_df, eval_df, summary)

    for qt in query_types:
        xlsx_path = queries_dir / f"queries_type_{qt}.xlsx"
        if not xlsx_path.exists():
            print(f"\n  WARNING: {xlsx_path.name} not found — skipping type {qt}")
            continue

        print("=" * 60)
        print(f"  Query Type {qt}")
        print("=" * 60)
        print(f"  Source: {xlsx_path.name}")

        queries = load_queries(xlsx_path)
        print(f"  Loaded {len(queries)} queries")

        t0 = time.time()
        retriever_results = run_all_retrievers(queries)

        for label in retriever_keys:
            total_hits = sum(len(v) for v in retriever_results[label].values())
            qids_with_hits = sum(1 for v in retriever_results[label].values() if v)
            print(f"    {label:<15} {total_hits:>6} hits across {qids_with_hits} queries")

        fused_df, eval_df, summary = process_query_type(
            query_type=qt,
            queries=queries,
            all_retriever_results=retriever_results,
            retriever_keys=retriever_keys,
        )
        summary["elapsed_s"] = round(time.time() - t0, 1)

        # Write per-type CSV
        csv_path = output_dir / f"type_{qt}_fused_all_retrievers_audit_{ts}.csv"
        fused_df.to_csv(csv_path, index=False)
        print(f"\n  Wrote {len(fused_df)} rows → {csv_path}")

        all_summaries.append(summary)
        all_type_data.append((qt, fused_df, eval_df, summary))
        print(f"  Completed type {qt} in {summary['elapsed_s']}s\n")

    # ── Cleanup dense retrievers ────────────────────────────────────────
    dense_retrievers.clear()
    try:
        import torch
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
    except ImportError:
        pass

    # ── Write combined Excel workbook ─────────────────────────────────
    print("─" * 60)
    print("Writing combined Excel workbook...")

    xlsx_path = output_dir / f"all_types_fused_audit_{ts}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:

        # Per-type sheets (3 sheets each)
        for qt, fused_df, eval_df, summary in all_type_data:
            write_type_sheets(writer, qt, fused_df, eval_df, summary, retriever_keys)

        # Overall Summary sheet (last)
        overall_rows = []
        for s in all_summaries:
            med = f"{s['median_rank']:.0f}" if s['median_rank'] is not None else "N/A"
            row = {
                "Query Type": f"Type {s['query_type']}",
                "Total Queries": s["total_queries"],
                "Found": s["found"],
                "Missed": s["missed"],
                "Recall (%)": f"{s['recall_pct']:.1f}%",
                "Median Fused Rank": med,
            }
            for r in retriever_keys:
                row[f"Found by {r}"] = f"{s['per_retriever'][r]}/{s['total_queries']}"
            overall_rows.append(row)

        overall_df = pd.DataFrame(overall_rows)
        overall_df.to_excel(writer, index=False, sheet_name="Overall_Summary")
        ws = writer.sheets["Overall_Summary"]
        style_header(ws, len(overall_df.columns))
        auto_width(ws, overall_df)
        ws.freeze_panes = "B2"

    print(f"  Wrote → {xlsx_path}")
    print(f"  Sheets: {3 * len(all_type_data)} type sheets + 1 Overall_Summary")

    # ── Final console summary ─────────────────────────────────────────
    print("\n" + "=" * 60)
    print("  OVERALL SUMMARY")
    print("=" * 60)
    print(f"  {'Type':<8} {'Queries':>8} {'Found':>8} {'Missed':>8} {'Recall':>8} {'Med Rank':>10} {'Time':>7}")
    print("  " + "-" * 55)
    for s in all_summaries:
        med = f"{s['median_rank']:.0f}" if s['median_rank'] is not None else "N/A"
        print(f"  Type {s['query_type']:<3} {s['total_queries']:>8} {s['found']:>8} "
              f"{s['missed']:>8} {s['recall_pct']:>7.1f}% {med:>10} {s['elapsed_s']:>6.1f}s")
    print("=" * 60)

    print("\n  Output files:")
    for f in sorted(output_dir.glob(f"*{ts}*")):
        size = f.stat().st_size / 1024
        print(f"    {f.name:<60} {size:>7.1f} KB")
    print()


if __name__ == "__main__":
    main()
