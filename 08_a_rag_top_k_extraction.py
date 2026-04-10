"""
08 — RAG Top-K Extraction

Extract the top-K fused ranks from all query types (1–6), enrich them with:
  1. query_text from the respective queries_type_{N}.xlsx files
  2. Full document data from documents.jsonl (flattened identifiers + metadata)

This produces the retrieval context that feeds into a RAG pipeline.

Usage
-----
    python scripts/08_rag_top_k_extraction.py

    # Custom top-K (default: 10)
    python scripts/08_rag_top_k_extraction.py --top-k 20

    # Custom fused audit file
    python scripts/08_rag_top_k_extraction.py --audit results/all_types_fused_audit_20260409_202603.xlsx

Inputs
------
    results/all_types_fused_audit_{timestamp}.xlsx   — from the Ali & Kieran script
    data/potential_queries/queries_type_{1-6}.xlsx    — for query_text
    data/json_format_data/full/documents.jsonl       — full 1.2M document corpus

Outputs (in results/)
---------------------
    (RAG)_top_{K}_fused_with_full_data_{timestamp}.csv
    (RAG)_top_{K}_fused_with_full_data_{timestamp}.xlsx
        Sheets: Type1, Type2, ..., Type6, All_Combined
"""

import argparse
import ast
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Locate project root ────────────────────────────────────────────────────

def find_root() -> Path:
    marker = "data/raw_data"
    for key in ("IR_PROJECT_ROOT", "VSCODE_WORKSPACE_FOLDER"):
        v = os.environ.get(key)
        if v and (Path(v) / marker).exists():
            return Path(v)
    for p in [Path.cwd()] + list(Path.cwd().parents):
        if (p / marker).exists():
            return p
    raise FileNotFoundError("Cannot find project root.")


ROOT = find_root()
sys.path.insert(0, str(ROOT))

RUN_TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M%S")

# ── Excel formatting ───────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_sheet(ws, dataframe):
    """Apply header styling, auto-width, freeze panes, and auto-filter."""
    ncols = len(dataframe.columns)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    for col_idx, col_name in enumerate(dataframe.columns, start=1):
        max_len = max(len(str(col_name)), dataframe[col_name].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="08 — RAG Top-K Extraction: extract top-K fused results "
                    "enriched with query text and full document data."
    )
    parser.add_argument(
        "--top-k", type=int, default=10,
        help="Number of top fused ranks to extract per query (default: 10)"
    )
    parser.add_argument(
        "--audit", type=str, default=None,
        help="Path to all_types_fused_audit_{ts}.xlsx. "
             "If not provided, uses the most recent one in results/."
    )
    parser.add_argument(
        "--docs", type=str, default=None,
        help="Path to documents.jsonl (default: data/json_format_data/full/documents.jsonl)"
    )
    parser.add_argument(
        "--query", type=str, default=None,
        help="Query suffix to select, e.g. '003' for Q1_003, Q2_003, etc. "
             "If not provided, runs ALL queries per type."
    )
    args = parser.parse_args()

    top_k = args.top_k
    query_suffix = args.query
    ts = RUN_TIMESTAMP
    results_dir = ROOT / "results"
    queries_dir = ROOT / "data" / "potential_queries"

    # ── Resolve audit file ────────────────────────────────────────────
    if args.audit:
        audit_path = Path(args.audit)
    else:
        # Find the most recent all_types_fused_audit_*.xlsx
        candidates = sorted(results_dir.glob("all_types_fused_audit_*.xlsx"))
        if not candidates:
            print("ERROR: No all_types_fused_audit_*.xlsx found in results/.")
            print("  Run the Ali & Kieran script first.")
            return 1
        audit_path = candidates[-1]

    if not audit_path.exists():
        print(f"ERROR: Audit file not found: {audit_path}")
        return 1

    # ── Resolve documents.jsonl ───────────────────────────────────────
    if args.docs:
        docs_path = Path(args.docs)
    else:
        docs_path = ROOT / "data" / "json_format_data" / "full" / "documents.jsonl"
        if not docs_path.exists():
            docs_path = ROOT / "data" / "json_format_data" / "subset_100k" / "documents.jsonl"

    if not docs_path.exists():
        print(f"ERROR: documents.jsonl not found at {docs_path}")
        return 1

    print("=" * 60)
    print("  08 — RAG Top-K Extraction")
    print("=" * 60)
    print(f"  Audit file  : {audit_path.name}")
    print(f"  Documents   : {docs_path}")
    print(f"  Top-K       : {top_k} (Q3/Q4), 3 (Q1/Q2/Q5/Q6)")
    print(f"  Query       : {f'_{query_suffix} only' if query_suffix else 'ALL queries'}")
    print(f"  Timestamp   : {ts}")
    print()

    t0 = time.time()

    # ── Step 1: Extract top-K fused ranks ─────────────────────────────
    print("Step 1: Extracting top-K fused ranks...")
    xls = pd.ExcelFile(audit_path)
    rrf_sheets = [s for s in xls.sheet_names if "RRF_Fused_Results" in s]

    # Q3 and Q4: top 10; all others: top 3
    TOP_K_FULL_TYPES = {"Type3", "Type4"}

    all_topk = []
    for sheet in rrf_sheets:
        df = pd.read_excel(audit_path, sheet_name=sheet)
        sheet_type = sheet.split("_RRF_")[0]
        k = top_k if sheet_type in TOP_K_FULL_TYPES else 3

        if query_suffix:
            # Extract type number from sheet (e.g. "Type3" -> "3")
            type_num = sheet_type.replace("Type", "")
            target_qid = f"Q{type_num}_{query_suffix}"
            if target_qid not in df["query_id"].values:
                print(f"  {sheet}: SKIPPED — {target_qid} not found")
                continue
            topk = df[(df["query_id"] == target_qid) & (df["fused_rank"] <= k)][
                ["query_id", "doc_id", "fused_rank", "rrf_score"]
            ].copy()
            label = f"{target_qid} only, top {k}"
        else:
            # All queries
            topk = df[df["fused_rank"] <= k][
                ["query_id", "doc_id", "fused_rank", "rrf_score"]
            ].copy()
            n_queries = topk["query_id"].nunique()
            label = f"all {n_queries} queries, top {k}"

        all_topk.append(topk)
        print(f"  {sheet}: {len(topk)} rows ({label})")

    topk_df = pd.concat(all_topk, ignore_index=True)
    print(f"  Total: {len(topk_df)} rows\n")

    # ── Step 2: Merge with query files for query_text ─────────────────
    print("Step 2: Merging with query Excel files for query_text...")
    query_texts = {}
    for qt in range(1, 7):
        qpath = queries_dir / f"queries_type_{qt}.xlsx"
        if not qpath.exists():
            continue
        qdf = pd.read_excel(qpath)
        for _, row in qdf.iterrows():
            qid = str(row["query_id"])
            raw = row["query_texts"]
            try:
                texts = ast.literal_eval(raw)
            except (ValueError, SyntaxError):
                texts = [str(raw)]
            if isinstance(texts, str):
                texts = [texts]
            query_texts[qid] = ", ".join(texts)

    topk_df["query_text"] = topk_df["query_id"].map(query_texts)
    matched = topk_df["query_text"].notna().sum()
    print(f"  Matched {matched}/{len(topk_df)} rows with query_text\n")

    # ── Step 3: Load matching docs from documents.jsonl ───────────────
    unique_doc_ids = set(topk_df["doc_id"].unique())
    print(f"Step 3: Loading {len(unique_doc_ids)} unique doc_ids from {docs_path.name}...")

    doc_data = {}
    with open(docs_path, encoding="utf-8") as f:
        for line in f:
            if not line.strip():
                continue
            doc = json.loads(line)
            did = doc.get("doc_id", "")
            if did in unique_doc_ids:
                # Flatten identifiers
                identifiers = doc.get("identifiers", {})
                flat_ids = {}
                for k, v in identifiers.items():
                    flat_ids[f"id_{k}"] = (
                        ", ".join(str(x) for x in v) if isinstance(v, list) else str(v)
                    )

                # Flatten metadata
                metadata = doc.get("metadata", {})
                flat_meta = {}
                for k, v in metadata.items():
                    flat_meta[f"meta_{k}"] = (
                        ", ".join(str(x) for x in v) if isinstance(v, list) else str(v)
                    )

                doc_data[did] = {
                    "caption": doc.get("caption", ""),
                    "schema": doc.get("schema", ""),
                    "text_blob": str(doc.get("text_blob", ""))[:500],
                    "embedding_text": str(doc.get("embedding_text", ""))[:500],
                    "first_seen": doc.get("first_seen", ""),
                    "last_seen": doc.get("last_seen", ""),
                    "token_count": len(doc.get("tokens", [])),
                    **flat_ids,
                    **flat_meta,
                }

                if len(doc_data) == len(unique_doc_ids):
                    break

    print(f"  Found {len(doc_data)}/{len(unique_doc_ids)} docs")
    missing = unique_doc_ids - set(doc_data.keys())
    if missing:
        print(f"  WARNING: {len(missing)} doc_ids not found in {docs_path.name}")
        for d in sorted(missing)[:10]:
            print(f"    - {d}")
        if len(missing) > 10:
            print(f"    ... and {len(missing) - 10} more")

    # ── Step 4: Merge document data ───────────────────────────────────
    print(f"\nStep 4: Merging document data...")
    doc_df = pd.DataFrame.from_dict(doc_data, orient="index")
    doc_df.index.name = "doc_id"
    doc_df = doc_df.reset_index()

    merged = topk_df.merge(doc_df, on="doc_id", how="left")
    print(f"  Final shape: {merged.shape} ({len(merged.columns)} columns)\n")

    # ── Step 5: Write outputs ─────────────────────────────────────────
    print("Step 5: Writing output files...")

    # CSV
    csv_path = results_dir / f"08_(RAG)_top_{top_k}_fused_with_full_data_{ts}.csv"
    merged.to_csv(csv_path, index=False)
    print(f"  CSV: {csv_path.name}")

    # Excel
    xlsx_path = results_dir / f"08_(RAG)_top_{top_k}_fused_with_full_data_{ts}.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for qt in range(1, 7):
            prefix = f"Q{qt}_"
            type_df = merged[merged["query_id"].str.startswith(prefix)].copy()
            if type_df.empty:
                continue
            sheet_name = f"Type{qt}"
            type_df.to_excel(writer, index=False, sheet_name=sheet_name)
            style_sheet(writer.sheets[sheet_name], type_df)
            print(f"    {sheet_name}: {len(type_df)} rows")

        # All Combined sheet
        merged.to_excel(writer, index=False, sheet_name="All_Combined")
        style_sheet(writer.sheets["All_Combined"], merged)

    print(f"  Excel: {xlsx_path.name}")

    elapsed = time.time() - t0
    print(f"\n{'=' * 60}")
    print(f"  Done in {elapsed:.1f}s")
    print(f"  {len(merged)} rows, {len(merged.columns)} columns")
    print(f"  {len(unique_doc_ids)} unique documents enriched from {docs_path.name}")
    print(f"{'=' * 60}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
